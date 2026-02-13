"""
Parse all onboarding test cases from Excel file with multiple sheets
"""

from openpyxl import load_workbook
import json

# Load workbook
wb = load_workbook('../testcases/CHIC_Onboarding_Questionnaire_Test_Cases.xlsx')

print("="*70)
print("PARSING ONBOARDING TEST CASES")
print("="*70)
print(f"Sheets found: {wb.sheetnames}\n")

all_test_cases = []

# Sheets to process (excluding Summary)
test_sheets = [
    'General Onboarding',
    'Step 1 - Body Type',
    'Step 2 - Highlight Areas',
    'Step 3 - Minimize Areas',
    'Step 4 - Favorite Colors',
    'Step 5 - Style Description'
]

for sheet_name in test_sheets:
    if sheet_name in wb.sheetnames:
        print(f"Processing: {sheet_name}...")
        ws = wb[sheet_name]

        headers = []
        sheet_tests = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # Header row
                headers = [cell if cell else f'Column_{j}' for j, cell in enumerate(row)]
                print(f"  Headers: {headers[:5]}...")
            else:
                if row[0]:  # If first column has value (Test ID)
                    test_case = {'Sheet': sheet_name}
                    for j, header in enumerate(headers):
                        if j < len(row):
                            test_case[header] = row[j] if row[j] else ''
                    sheet_tests.append(test_case)

        all_test_cases.extend(sheet_tests)
        print(f"  Found {len(sheet_tests)} test cases\n")

# Save all test cases
with open('onboarding_test_cases.json', 'w', encoding='utf-8') as f:
    json.dump(all_test_cases, f, indent=2, ensure_ascii=False)

print("="*70)
print(f"TOTAL TEST CASES EXTRACTED: {len(all_test_cases)}")
print("="*70)

# Breakdown by sheet
print("\nBreakdown by Sheet:")
for sheet_name in test_sheets:
    count = sum(1 for tc in all_test_cases if tc.get('Sheet') == sheet_name)
    print(f"  {sheet_name}: {count} tests")

# Check for priorities if they exist
priorities = {}
for tc in all_test_cases:
    priority = tc.get('Priority', 'Unknown')
    if priority != '':
        priorities[priority] = priorities.get(priority, 0) + 1

if priorities:
    print("\nBy Priority:")
    for priority, count in sorted(priorities.items()):
        print(f"  {priority}: {count}")

# Check for types
types = {}
for tc in all_test_cases:
    test_type = tc.get('Test Type', 'Unknown')
    if test_type != '':
        types[test_type] = types.get(test_type, 0) + 1

if types:
    print("\nBy Type:")
    for test_type, count in sorted(types.items()):
        print(f"  {test_type}: {count}")

# Show first 5 test cases
print("\nFirst 5 Test Cases:")
for i, tc in enumerate(all_test_cases[:5]):
    test_id = tc.get('Test Case ID', tc.get('TC ID', 'Unknown'))
    test_desc = tc.get('Test Description', tc.get('Description', 'No description'))
    print(f"  {test_id}: {test_desc[:60]}...")

print("\n" + "="*70)
print("Saved to: onboarding_test_cases.json")
print("="*70 + "\n")
