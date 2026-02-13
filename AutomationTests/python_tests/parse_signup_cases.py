from openpyxl import load_workbook
import json

# Load workbook
wb = load_workbook('../testcases/CHIC_SignUp_Page_Test_Cases.xlsx')
ws = wb['Sign Up Page Test Cases']

# Parse test cases
test_cases = []
headers = []

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:  # Header row
        headers = [cell if cell else f'Column_{j}' for j, cell in enumerate(row)]
    else:
        if row[0]:  # If Test Case ID exists
            test_case = {}
            for j, header in enumerate(headers):
                test_case[header] = row[j] if row[j] else ''
            test_cases.append(test_case)

# Save as JSON
with open('signup_test_cases.json', 'w', encoding='utf-8') as f:
    json.dump(test_cases, f, indent=2, ensure_ascii=False)

# Print summary
print(f"Total test cases extracted: {len(test_cases)}")
print("\nTest Cases by Priority:")
priorities = {}
for tc in test_cases:
    priority = tc.get('Priority', 'Unknown')
    priorities[priority] = priorities.get(priority, 0) + 1

for priority, count in sorted(priorities.items()):
    print(f"  {priority}: {count}")

print("\nTest Cases by Type:")
types = {}
for tc in test_cases:
    test_type = tc.get('Test Type', 'Unknown')
    types[test_type] = types.get(test_type, 0) + 1

for test_type, count in sorted(types.items()):
    print(f"  {test_type}: {count}")

# Print first 5 test case IDs
print("\nFirst 10 Test Cases:")
for i, tc in enumerate(test_cases[:10]):
    print(f"  {tc['Test Case ID']}: {tc['Test Description'][:60]}...")
